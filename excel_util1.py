import openpyxl
def get_excel_data(file_path,sheet_name):
    #open the workbook
    workbook = openpyxl.load_workbook(file_path)
    #fetch the sheet name
    sheet = workbook[sheet_name]
    data =[]
    for row in range(2, sheet.max_row + 1):
        username = sheet.cell(row=row,column=1).value
        password = sheet.cell(row=row, column=2).value
        firstname = sheet.cell(row=row,column=3).value
        lastname=sheet.cell(row=row,column=4).value
        postalcode=sheet.cell(row=row,column=5).value
        data.append((username,password,firstname,lastname,postalcode))
    return data
