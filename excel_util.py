import openpyxl


def get_excel_data1(file_path, sheet_name):
    # open the work book
    workbook = openpyxl.load_workbook(file_path)
    # fetch  the sheet name
    sheet = workbook[sheet_name]

    data = []

    #for row in range(2, sheet.max_row+1 ):
        #username = sheet.cell(row=row, column=1).value
        #password = sheet.cell(row=row, column=2).value
        #fname = sheet.cell(row=row, column=3).value
        #lname = sheet.cell(row=row, column=4).value
        #zcode = sheet.cell(row=row, column=5).value
        #data.append((username, password))

    for row in range(2, sheet.max_row+1 ):
        First_name = sheet.cell(row=row, column=1).value
        Last_name = sheet.cell(row=row, column=2).value
        Post = sheet.cell(row=row, column=3).value
        Money = sheet.cell(row=row, column=4).value
        Full_name = sheet.cell(row=row, column=5).value
        Deposit = sheet.cell(row=row, column=6).value
        withdraw = sheet.cell(row=row, column=7).value
        data.append((First_name, Last_name,Post ,Money,Full_name,Deposit,withdraw))

    return data